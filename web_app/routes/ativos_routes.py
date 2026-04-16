from __future__ import annotations

import csv
import hashlib
import io
import json
import time
from datetime import datetime

from flask import flash, jsonify, redirect, render_template, request, send_file, session, url_for
from services.storage_backend import S3StorageBackend, StorageBackendError
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from models.ativos import Ativo
from services.ativos_arquivo_service import (
    ArquivoInvalido,
    ArquivoNaoEncontrado,
    TipoDocumentoInvalido,
)
from services.ativos_service import AtivoErro, AtivoNaoEncontrado, AtivosService, PermissaoNegada
from utils.auth import require_auth_api
from utils.csrf import validar_csrf_da_requisicao, require_csrf
from utils.validators import STATUS_VALIDOS, SETORES_VALIDOS, CONDICOES_VALIDAS, UNIDADES_VALIDAS, TIPOS_ATIVO_VALIDOS


# Mensagens padronizadas para manter respostas da camada web consistentes.
MSG_SESSAO_EXPIRADA = "Sessão expirada. Faça login novamente."
MSG_ERRO_LISTAR_ATIVOS = "Erro inesperado ao listar ativos."

# ========== PROTEÇÃO CONTRA DUPLICAÇÃO DE CADASTRO ==========
# Cache em memória para rastrear requisições de criação recentes.
# Evita duplicação quando usuário clica múltiplas vezes em "Confirmar".
# Chave: hash(dados_do_ativo + user_id), Valor: (ativo_id, timestamp)
# Expire automático: requisições com >10 segundos são ignoradas (nova criação)
_creation_dedup_cache = {}

def _gerar_chave_dedup(dados: dict, user_id: int) -> str:
    """
    Gera uma chave única (hash) para deduplicação baseada nos dados principais do ativo.
    Ignora campos que mudam automaticamente (id, criado_em, atualizado_em).
    """
    # Campos utilizados para deduplicação.
    # Se esses campos forem iguais, consideramos a criação como duplicada.
    chaves_dedup = [
        "tipo", "tipo_ativo", "marca", "modelo", "serial",
        "usuario_responsavel", "setor", "departamento", "localizacao",
        "condicao", "status"
    ]

    # Extrai apenas os campos relevantes e cria uma string normalizada.
    dados_norm = {k: str(dados.get(k, "")).strip().lower() for k in chaves_dedup}

    # Concatena dados + user_id e gera hash SHA256.
    chave_str = f"{json.dumps(dados_norm, sort_keys=True)}#{user_id}"
    return hashlib.sha256(chave_str.encode()).hexdigest()

def _verificar_duplicacao(chave_dedup: str, ativo_id_novo: str) -> tuple[bool, str | None]:
    """
    Verifica se uma criação é duplicada (requisição recente com mesmos dados).
    Se sim, retorna (True, id_do_ativo_existente).
    Se não, registra esta criação e retorna (False, None).

    Limpeza automática: remove entradas com >10 segundos.
    """
    tempo_atual = time.time()

    # Limpa cache: remove entradas antigas.
    chaves_expiradas = [
        k for k, (_, timestamp) in _creation_dedup_cache.items()
        if tempo_atual - timestamp > 10
    ]
    for k in chaves_expiradas:
        del _creation_dedup_cache[k]

    # Verifica se esta chave foi criada recentemente.
    if chave_dedup in _creation_dedup_cache:
        ativo_id_existente, _ = _creation_dedup_cache[chave_dedup]

        # Primeiro request pode registrar a chave sem ID; quando o ID for conhecido,
        # atualiza o cache para que replays posteriores retornem o mesmo ativo.
        if not ativo_id_existente and ativo_id_novo:
            _creation_dedup_cache[chave_dedup] = (ativo_id_novo, tempo_atual)
            return (False, None)

        # Duplicação detectada — retorna ativo já criado.
        return (True, ativo_id_existente)

    # Nova criação — registra no cache.
    _creation_dedup_cache[chave_dedup] = (ativo_id_novo, tempo_atual)
    return (False, None)


def _obter_user_id_logado() -> int | None:
    """
    Obtém o identificador do usuário autenticado na sessão.
    """
    user_id = session.get("user_id")
    if user_id is None:
        return None
    return int(user_id)


def _request_data() -> dict:
    """
    Obtém o payload JSON ou o form-data da requisição.
    """
    if request.is_json:
        return request.get_json(silent=True) or {}
    return request.form.to_dict()


def _json_success(message: str, *, status: int = 200, **payload):
    """
    Padroniza respostas JSON de sucesso.
    """
    body = {"ok": True, "mensagem": message}
    body.update(payload)
    return jsonify(body), status


def _json_error(message: str, status: int = 400, **payload):
    """
    Padroniza respostas JSON de erro.
    """
    body = {"ok": False, "erro": message}
    body.update(payload)
    return jsonify(body), status


def _normalizar_flag_presenca(raw_value: str | None) -> str | None:
    """
    Normaliza flags de presença recebidas no filtro (sim/nao).
    """
    value = (raw_value or "").strip().lower()
    if value in {"sim", "true", "1", "com"}:
        return "sim"
    if value in {"nao", "false", "0", "sem"}:
        return "nao"
    return None


def _serializar_ativo(ativo: Ativo) -> dict:
    """
    Serializa o ativo no contrato mínimo esperado pelo dashboard.
    """
    return {
        "id": ativo.id_ativo,
        # Expõe tipo e tipo_ativo para preservar compatibilidade com telas legadas.
        "tipo": ativo.tipo,
        "tipo_ativo": getattr(ativo, "tipo_ativo", None) or ativo.tipo,
        "marca": ativo.marca,
        "modelo": ativo.modelo,
        "serial": getattr(ativo, "serial", "") or "",
        "codigo_interno": getattr(ativo, "codigo_interno", "") or "",
        "descricao": getattr(ativo, "descricao", "") or "",
        "categoria": getattr(ativo, "categoria", "") or "",
        "condicao": getattr(ativo, "condicao", "") or "",
        "localizacao": getattr(ativo, "localizacao", "") or "",
        "setor": getattr(ativo, "setor", "") or getattr(ativo, "departamento", "") or "",
        "usuario_responsavel": ativo.usuario_responsavel or "",
        "email_responsavel": getattr(ativo, "email_responsavel", "") or "",
        "departamento": ativo.departamento or "",
        "status": ativo.status or "",
        "data_entrada": ativo.data_entrada or "",
        "data_saida": ativo.data_saida or "",
        "data_compra": getattr(ativo, "data_compra", "") or "",
        "valor": getattr(ativo, "valor", "") or "",
        "observacoes": getattr(ativo, "observacoes", "") or "",
        "detalhes_tecnicos": getattr(ativo, "detalhes_tecnicos", "") or "",
        "processador": getattr(ativo, "processador", "") or "",
        "ram": getattr(ativo, "ram", "") or "",
        "armazenamento": getattr(ativo, "armazenamento", "") or "",
        "sistema_operacional": getattr(ativo, "sistema_operacional", "") or "",
        "carregador": getattr(ativo, "carregador", "") or "",
        "teamviewer_id": getattr(ativo, "teamviewer_id", "") or "",
        "anydesk_id": getattr(ativo, "anydesk_id", "") or "",
        "nome_equipamento": getattr(ativo, "nome_equipamento", "") or "",
        "hostname": getattr(ativo, "hostname", "") or "",
        "imei_1": getattr(ativo, "imei_1", "") or "",
        "imei_2": getattr(ativo, "imei_2", "") or "",
        "numero_linha": getattr(ativo, "numero_linha", "") or "",
        "operadora": getattr(ativo, "operadora", "") or "",
        "conta_vinculada": getattr(ativo, "conta_vinculada", "") or "",
        "polegadas": getattr(ativo, "polegadas", "") or "",
        "resolucao": getattr(ativo, "resolucao", "") or "",
        "tipo_painel": getattr(ativo, "tipo_painel", "") or "",
        "entrada_video": getattr(ativo, "entrada_video", "") or "",
        "fonte_ou_cabo": getattr(ativo, "fonte_ou_cabo", "") or "",
        "created_at": getattr(ativo, "created_at", None) or getattr(ativo, "criado_em", None) or "",
        "updated_at": getattr(ativo, "updated_at", None) or getattr(ativo, "atualizado_em", None) or "",
        "data_ultima_movimentacao": getattr(ativo, "data_ultima_movimentacao", None) or "",
        # Usa getattr para manter compatibilidade com objetos de teste simplificados.
        "nota_fiscal": getattr(ativo, "nota_fiscal", "") or "",
        "garantia": getattr(ativo, "garantia", "") or "",
        "resumo_movimentacao": getattr(ativo, "resumo_movimentacao", None),
    }


def _serializar_arquivo(arquivo: dict) -> dict:
    """
    Serializa um registro de arquivo para o frontend.
    """
    return {
        "id": arquivo["id"],
        "ativo_id": arquivo["ativo_id"],
        "tipo_documento": arquivo["tipo_documento"],
        "nome_original": arquivo["nome_original"],
        "tamanho_bytes": arquivo["tamanho_bytes"],
        "mime_type": arquivo["mime_type"],
        "criado_em": arquivo["criado_em"],
    }


def _eh_admin(perfil: str | None) -> bool:
    """
    Centraliza a regra de detecção de perfil administrativo.
    Verificação de visibilidade de campos técnicos/sensíveis.
    """
    return (perfil or "").strip().lower() in {"adm", "admin"}


def _resumo_ativo_para_modal(ativo: dict, eh_admin: bool) -> dict:
    """
    Retorna um resumo composto do ativo para exibição em modal.
    Mostra informações principais e opcionalmente campos técnicos conforme tipo.

    Campos sempre visíveis:
    - ID, tipo, marca, modelo, status
    - responsável, e-mail, setor/departamento
    - data de entrada, data de saída
    - presença de nota fiscal e garantia

    Campos específicos por tipo (quando aplicável):
    - Notebook/Desktop: processador, RAM, armazenamento, sistema operacional
    - Monitor: polegadas, resolução, tipo de painel, entradas de vídeo
    - Celular: IMEI principal, número da linha, operadora

    Campos técnicos restritos (visíveis apenas para admin):
    - AnyDesk ID, TeamViewer ID, hostname, serial, código interno
    """
    tipo = (ativo.get("tipo") or "").strip().lower()

    # Seção principal — sempre visível
    resumo = {
        "secao_principal": {
            "id": ativo.get("id", ""),
            "tipo": ativo.get("tipo", ""),
            "marca": ativo.get("marca", ""),
            "modelo": ativo.get("modelo", ""),
            "status": ativo.get("status", ""),
        },
        "secao_responsabilidade": {
            "usuario_responsavel": ativo.get("usuario_responsavel", ""),
            "email_responsavel": ativo.get("email_responsavel", ""),
            "setor": ativo.get("setor", ""),
            "localizacao": ativo.get("localizacao", ""),
        },
        "secao_ciclo": {
            "data_entrada": ativo.get("data_entrada", ""),
            "data_saida": ativo.get("data_saida", ""),
            "nota_fiscal": ativo.get("nota_fiscal", ""),
            "garantia": ativo.get("garantia", ""),
        },
    }

    # Seção técnica por tipo — visível para todos quando não restritos
    resumo["secao_tecnica"] = {}

    if tipo in {"notebook", "desktop"}:
        resumo["secao_tecnica"]["label"] = "Especificações de Computador"
        resumo["secao_tecnica"]["campos"] = {
            "processador": ativo.get("processador", ""),
            "ram": ativo.get("ram", ""),
            "armazenamento": ativo.get("armazenamento", ""),
            "sistema_operacional": ativo.get("sistema_operacional", ""),
        }
    elif tipo == "monitor":
        resumo["secao_tecnica"]["label"] = "Especificações de Monitor"
        resumo["secao_tecnica"]["campos"] = {
            "polegadas": ativo.get("polegadas", ""),
            "resolucao": ativo.get("resolucao", ""),
            "tipo_painel": ativo.get("tipo_painel", ""),
            "entrada_video": ativo.get("entrada_video", ""),
        }
    elif tipo == "celular":
        resumo["secao_tecnica"]["label"] = "Especificações de Celular"
        resumo["secao_tecnica"]["campos"] = {
            "imei_1": ativo.get("imei_1", ""),
            "numero_linha": ativo.get("numero_linha", ""),
            "operadora": ativo.get("operadora", ""),
        }

    # Seção restrita — visível apenas para admin
    if eh_admin:
        resumo["secao_tecnica_restrita"] = {
            "anydesk_id": ativo.get("anydesk_id", ""),
            "teamviewer_id": ativo.get("teamviewer_id", ""),
            "hostname": ativo.get("hostname", ""),
            "serial": ativo.get("serial", ""),
            "codigo_interno": ativo.get("codigo_interno", ""),
        }

    return resumo


def _mapa_campos_ativo(dados: dict) -> dict:
    """
    Centraliza o mapeamento normalizado de campos de um payload de ativo.

    Esse mapa é usado tanto na criação (para construir Ativo) quanto na atualização
    (para normalizar payload antes de chamar o service). Centralizar evita duplicação
    e facilita manutenção futura quando o schema mudar.

    Prioriza:
    - tipo_ativo sobre tipo (compatibilidade com chave legada)
    - setor sobre departamento (compatibilidade com chave legada)
    """
    return {
        "tipo_ativo": dados.get("tipo_ativo", dados.get("tipo", "")),
        "marca": dados.get("marca", ""),
        "modelo": dados.get("modelo", ""),
        "serial": dados.get("serial", "") or None,
        "codigo_interno": dados.get("codigo_interno", "") or None,
        "descricao": dados.get("descricao", ""),
        "categoria": dados.get("categoria", ""),
        "condicao": dados.get("condicao", "") or None,
        "localizacao": dados.get("localizacao", "") or None,
        "setor": dados.get("setor", dados.get("departamento", "")),
        "usuario_responsavel": dados.get("usuario_responsavel", "") or None,
        "email_responsavel": dados.get("email_responsavel", "") or None,
        "departamento": dados.get("setor", dados.get("departamento", "")),
        "status": dados.get("status", ""),
        "data_entrada": dados.get("data_entrada", ""),
        "data_saida": dados.get("data_saida", "") or None,
        "data_compra": dados.get("data_compra", "") or None,
        "valor": dados.get("valor", "") or None,
        "observacoes": dados.get("observacoes", "") or None,
        "detalhes_tecnicos": dados.get("detalhes_tecnicos", "") or None,
        "processador": dados.get("processador", "") or None,
        "ram": dados.get("ram", "") or None,
        "armazenamento": dados.get("armazenamento", "") or None,
        "sistema_operacional": dados.get("sistema_operacional", "") or None,
        "carregador": dados.get("carregador", "") or None,
        "teamviewer_id": dados.get("teamviewer_id", "") or None,
        "anydesk_id": dados.get("anydesk_id", "") or None,
        "nome_equipamento": dados.get("nome_equipamento", "") or None,
        "hostname": dados.get("hostname", "") or None,
        "imei_1": dados.get("imei_1", "") or None,
        "imei_2": dados.get("imei_2", "") or None,
        "numero_linha": dados.get("numero_linha", "") or None,
        "operadora": dados.get("operadora", "") or None,
        "conta_vinculada": dados.get("conta_vinculada", "") or None,
        "polegadas": dados.get("polegadas", "") or None,
        "resolucao": dados.get("resolucao", "") or None,
        "tipo_painel": dados.get("tipo_painel", "") or None,
        "entrada_video": dados.get("entrada_video", "") or None,
        "fonte_ou_cabo": dados.get("fonte_ou_cabo", "") or None,
        "nota_fiscal": None,
        "garantia": None,
    }


def _ativo_do_payload(dados: dict) -> Ativo:
    """
    Constrói o domínio Ativo a partir do payload do frontend.
    O campo id não é lido do payload — o ID é gerado automaticamente no backend.

    Usa _mapa_campos_ativo() para normalizar campos, evitando duplicação com
    _normalizar_payload_atualizacao().
    """
    # Obtém o mapa normalizado de campos
    mapa = _mapa_campos_ativo(dados)

    # Constrói o Ativo com os campos normalizados
    return Ativo(
        id_ativo=None,  # gerado pelo service no momento do cadastro
        tipo=mapa["tipo_ativo"],  # campo tipo é alias de tipo_ativo no Ativo
        marca=mapa["marca"],
        modelo=mapa["modelo"],
        serial=mapa["serial"],
        codigo_interno=mapa["codigo_interno"],
        descricao=mapa["descricao"],
        categoria=mapa["categoria"],
        tipo_ativo=mapa["tipo_ativo"],
        condicao=mapa["condicao"],
        localizacao=mapa["localizacao"],
        setor=mapa["setor"],
        usuario_responsavel=mapa["usuario_responsavel"],
        email_responsavel=mapa["email_responsavel"],
        departamento=mapa["departamento"],
        status=mapa["status"],
        data_entrada=mapa["data_entrada"],
        data_saida=mapa["data_saida"],
        data_compra=mapa["data_compra"],
        valor=mapa["valor"],
        observacoes=mapa["observacoes"],
        detalhes_tecnicos=mapa["detalhes_tecnicos"],
        processador=mapa["processador"],
        ram=mapa["ram"],
        armazenamento=mapa["armazenamento"],
        sistema_operacional=mapa["sistema_operacional"],
        carregador=mapa["carregador"],
        teamviewer_id=mapa["teamviewer_id"],
        anydesk_id=mapa["anydesk_id"],
        nome_equipamento=mapa["nome_equipamento"],
        hostname=mapa["hostname"],
        imei_1=mapa["imei_1"],
        imei_2=mapa["imei_2"],
        numero_linha=mapa["numero_linha"],
        operadora=mapa["operadora"],
        conta_vinculada=mapa["conta_vinculada"],
        polegadas=mapa["polegadas"],
        resolucao=mapa["resolucao"],
        tipo_painel=mapa["tipo_painel"],
        entrada_video=mapa["entrada_video"],
        fonte_ou_cabo=mapa["fonte_ou_cabo"],
        nota_fiscal=mapa["nota_fiscal"],
        garantia=mapa["garantia"],
    )


def _normalizar_payload_atualizacao(dados: dict, *, preencher_campos_ausentes: bool = True) -> dict:
    """
    Normaliza o payload de edição em um contrato único para PUT, prévia e confirmação.

    Usa _mapa_campos_ativo() para obter o mapeamento padrão, evitando duplicação com
    _ativo_do_payload(). O parâmetro preencher_campos_ausentes controla se a função
    retorna todos os campos (com defaults) ou apenas os campos presentes no payload.

    Parâmetros:
      dados: dict — payload bruto do frontend ou JavaScript
      preencher_campos_ausentes: bool — se True, retorna mapa completo com defaults;
                                        se False, retorna apenas campos enviados pelo front
    """
    # Obtém o mapa completo de campos normalizados
    mapa_campos = _mapa_campos_ativo(dados)

    # Se preencher_campos_ausentes=True, retorna o mapa completo com defaults
    if preencher_campos_ausentes:
        return mapa_campos

    # Caso contrário, retorna apenas os campos que foram realmente enviados no payload
    dados_normalizados: dict = {}
    for chave, valor in mapa_campos.items():
        # Exclui campos que não vêm do payload (nota_fiscal e garantia são sempre None)
        if chave in {"nota_fiscal", "garantia"}:
            continue
        # Inclui apenas se a chave foi enviada no payload original
        if chave in dados:
            dados_normalizados[chave] = valor

    # Mantém compatibilidade de chave legada apenas quando o front realmente enviar dado operacional.
    if "setor" in dados_normalizados and "departamento" not in dados_normalizados:
        dados_normalizados["departamento"] = dados_normalizados["setor"]

    return dados_normalizados


def _validar_intervalo_datas(data_inicial: str | None, data_final: str | None, *, campo: str) -> None:
    """
    Valida consistencia de intervalo de datas de filtros (inicial <= final).
    """
    if not data_inicial or not data_final:
        return

    try:
        inicio = datetime.strptime(data_inicial, "%Y-%m-%d")
        fim = datetime.strptime(data_final, "%Y-%m-%d")
    except ValueError as exc:
        raise AtivoErro(f"Intervalo de {campo} invalido.") from exc

    if inicio > fim:
        raise AtivoErro(f"Data inicial de {campo} nao pode ser maior que a data final.")


def _linhas_exportacao(ativos: list[Ativo]) -> list[dict]:
    """
    Padroniza o dataset de exportacao para todos os formatos.
    """
    return [
        {
            "id": ativo.id_ativo,
            "tipo": ativo.tipo or "",
            "marca": ativo.marca or "",
            "modelo": ativo.modelo or "",
            "usuario_responsavel": ativo.usuario_responsavel or "",
            "departamento": ativo.departamento or "",
            "status": ativo.status or "",
            "data_entrada": ativo.data_entrada or "",
            "data_saida": ativo.data_saida or "",
            "nota_fiscal": getattr(ativo, "nota_fiscal", "") or "",
            "garantia": getattr(ativo, "garantia", "") or "",
        }
        for ativo in ativos
    ]


def _nome_arquivo_exportacao(formato: str) -> str:
    """
    Gera nome de arquivo amigavel e deterministico para downloads.
    """
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"ativos_export_{timestamp}.{formato}"


def _resposta_sem_registros_exportacao():
    """
    Mantem resposta consistente quando nao existem registros para exportar.
    """
    return _json_error("Nao existem ativos para exportacao com os filtros informados.", status=404)


def _status_documento_exportacao(valor_documento: str) -> str:
    """
    Retorna texto curto e profissional para colunas documentais no relatorio.
    """
    return "Vinculada" if (valor_documento or "").strip() else ""


def _texto_curto_pdf(valor: str, limite: int = 28) -> str:
    """
    Evita estouro visual no PDF sem perder o sentido do campo.
    """
    texto = (valor or "").strip()
    if len(texto) <= limite:
        return texto
    return f"{texto[:limite - 3]}..."


def _gerar_xlsx_em_memoria(linhas: list[dict]) -> io.BytesIO:
    """
    Monta o arquivo XLSX em memoria com cabecalho legivel e ajuste de colunas.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ativos"

    colunas = [
        ("id", "ID"),
        ("tipo", "Tipo"),
        ("marca", "Marca"),
        ("modelo", "Modelo"),
        ("usuario_responsavel", "Responsavel"),
        ("departamento", "Departamento"),
        ("status", "Status"),
        ("data_entrada", "Data Entrada"),
        ("data_saida", "Data Saida"),
        ("nota_fiscal", "Nota Fiscal"),
        ("garantia", "Garantia"),
    ]

    worksheet.append([titulo for _, titulo in colunas])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for linha in linhas:
        worksheet.append(
            [
                linha["id"],
                linha["tipo"],
                linha["marca"],
                linha["modelo"],
                linha["usuario_responsavel"],
                linha["departamento"],
                linha["status"],
                linha["data_entrada"],
                linha["data_saida"],
                _status_documento_exportacao(linha["nota_fiscal"]),
                _status_documento_exportacao(linha["garantia"]),
            ]
        )

    # Mantem layout legivel sem deixar nomes tecnicos dominarem a planilha.
    larguras = {
        "A": 14,
        "B": 14,
        "C": 14,
        "D": 16,
        "E": 20,
        "F": 16,
        "G": 12,
        "H": 14,
        "I": 14,
        "J": 12,
        "K": 12,
    }
    for coluna, largura in larguras.items():
        worksheet.column_dimensions[coluna].width = largura

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Guarda o nome completo do arquivo em comentario sem poluir a coluna principal.
    for row_index, linha in enumerate(linhas, start=2):
        if (linha.get("nota_fiscal") or "").strip():
            worksheet.cell(row=row_index, column=10).comment = Comment(
                f"Arquivo vinculado: {linha['nota_fiscal']}",
                "Sistema"
            )
        if (linha.get("garantia") or "").strip():
            worksheet.cell(row=row_index, column=11).comment = Comment(
                f"Arquivo vinculado: {linha['garantia']}",
                "Sistema"
            )

    xlsx_bytes = io.BytesIO()
    workbook.save(xlsx_bytes)
    xlsx_bytes.seek(0)
    return xlsx_bytes


def _gerar_pdf_em_memoria(linhas: list[dict]) -> io.BytesIO:
    """
    Gera PDF tabular em memoria para exportacao de ativos.
    """
    pdf_buffer = io.BytesIO()
    documento = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
        title="Relatorio de Ativos",
    )

    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph("Relatorio de Ativos", estilos["Title"]),
        Spacer(1, 8),
        Paragraph(
            f"Gerado em: {datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S')} UTC",
            estilos["Normal"],
        ),
        Spacer(1, 12),
    ]

    cabecalho = [
        "ID",
        "Tipo",
        "Marca",
        "Modelo",
        "Responsavel",
        "Departamento",
        "Status",
        "Entrada",
        "Saida",
        "Nota Fiscal",
        "Garantia",
    ]
    dados_tabela = [cabecalho]

    for linha in linhas:
        dados_tabela.append(
            [
                _texto_curto_pdf(str(linha["id"]), limite=16),
                _texto_curto_pdf(str(linha["tipo"]), limite=14),
                _texto_curto_pdf(str(linha["marca"]), limite=14),
                _texto_curto_pdf(str(linha["modelo"]), limite=16),
                _texto_curto_pdf(str(linha["usuario_responsavel"]), limite=20),
                _texto_curto_pdf(str(linha["departamento"]), limite=16),
                _texto_curto_pdf(str(linha["status"]), limite=12),
                str(linha["data_entrada"]),
                str(linha["data_saida"]),
                _status_documento_exportacao(str(linha["nota_fiscal"])),
                _status_documento_exportacao(str(linha["garantia"])),
            ]
        )

    tabela = Table(
        dados_tabela,
        repeatRows=1,
        colWidths=[56, 60, 60, 64, 78, 70, 52, 58, 58, 62, 62],
    )
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f1020")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#7a7a7a")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    elementos.append(tabela)
    documento.build(elementos)
    pdf_buffer.seek(0)
    return pdf_buffer


def registrar_rotas_ativos(app, *, ativos_service: AtivosService, ativos_arquivo_service):
    """
    Registra a camada HTTP do dashboard e do CRUD mínimo de ativos.
    """
    service = ativos_service
    arquivo_service = ativos_arquivo_service

    # DEPRECATED: Esta função helper foi substituída pelo decorator @require_csrf()
    # Mantida apenas para referência histórica — remover em próxima rodada de refatoração.
    # Veja utils/csrf.py::require_csrf() para a implementação moderna.
    def _validar_csrf_endpoints_mutacao():
        """
        [DEPRECATED] Proteção CSRF manual — use @require_csrf() decorator em vez disso.

        Razão: O decorator @require_csrf() encapsula melhor a validação CSRF,
        evita duplicação de código nas rotas, e torna explícita a intenção de
        segurança ao olhar a definição da rota.
        """
        if validar_csrf_da_requisicao(request):
            return None
        return _json_error(
            "Requisição inválida. Atualize a página e tente novamente.",
            status=403,
        )

    def _resolver_documento_vinculado(arquivos: list[dict], tipo_documento: str) -> str:
        """
        Resolve o documento vinculado por categoria usando o primeiro registro valido.
        Regra adotada: usar o primeiro da lista retornada pelo service (ordem decrescente de criacao).
        """
        for arquivo in arquivos:
            if arquivo.get("tipo_documento") != tipo_documento:
                continue
            nome_original = (arquivo.get("nome_original") or "").strip()
            if nome_original:
                return nome_original
        return ""

    def _linhas_exportacao_enriquecidas(ativos: list[Ativo], user_id: int) -> list[dict]:
        """
        Monta linhas de exportacao refletindo anexos reais (nota fiscal/garantia),
        com fallback para campos legados do ativo quando nao houver anexo da categoria.
        """
        linhas = _linhas_exportacao(ativos)

        for ativo, linha in zip(ativos, linhas):
            ativo_id = str(ativo.id_ativo)
            arquivos = arquivo_service.listar_arquivos(ativo_id, user_id)

            nota_fiscal_vinculada = _resolver_documento_vinculado(arquivos, "nota_fiscal")
            garantia_vinculada = _resolver_documento_vinculado(arquivos, "garantia")

            # Mantem compatibilidade com campos legados quando nao houver anexo vinculado.
            linha["nota_fiscal"] = nota_fiscal_vinculada or linha.get("nota_fiscal", "")
            linha["garantia"] = garantia_vinculada or linha.get("garantia", "")

        return linhas

    def _coletar_filtros_e_ordenacao_da_query() -> tuple[dict, str, str, str | None, str | None]:
        """
        Centraliza leitura de query params para manter consistência entre lista e exportação.
        Coleta: id, tipo, marca, modelo, responsável, departamento, localidade, status,
        presença de documentos (garantia/nota fiscal), períodos de data entrada/saída.
        """
        # Coleta todos os filtros suportados a partir da query string.
        # Valores vazios são normalizados para None para manter dict limpo.
        filtros = {
            "id_ativo": request.args.get("id_ativo", "").strip() or None,
            "tipo": request.args.get("tipo", "").strip() or None,
            "marca": request.args.get("marca", "").strip() or None,
            "modelo": request.args.get("modelo", "").strip() or None,
            "usuario_responsavel": request.args.get("usuario_responsavel", "").strip() or None,
            "departamento": request.args.get("departamento", "").strip() or None,
            "localizacao": request.args.get("localizacao", "").strip() or None,
            "nota_fiscal": request.args.get("nota_fiscal", "").strip() or None,
            "garantia": request.args.get("garantia", "").strip() or None,
            "status": request.args.get("status", "").strip() or None,
            "data_entrada_inicial": request.args.get("data_entrada_inicial", "").strip() or None,
            "data_entrada_final": request.args.get("data_entrada_final", "").strip() or None,
            "data_saida_inicial": request.args.get("data_saida_inicial", "").strip() or None,
            "data_saida_final": request.args.get("data_saida_final", "").strip() or None,
        }
        # Remove entradas nulas para manter dict apenas com filtros realmente solicitados.
        filtros = {k: v for k, v in filtros.items() if v is not None}

        ordenar_por = request.args.get("ordenar_por", "id").strip() or "id"
        ordem = request.args.get("ordem", "asc").strip().lower() or "asc"

        if ordem not in {"asc", "desc"}:
            raise AtivoErro("Ordem de classificacao invalida. Use asc ou desc.")

        # Lista de campos permitidos para ordenação da listagem de ativos.
        # Apenas esses campos podem ser usados para ORDER BY para evitar injeção.
        campos_ordenacao_permitidos = {
            "id",
            "tipo",
            "marca",
            "modelo",
            "usuario_responsavel",
            "departamento",
            "localizacao",
            "nota_fiscal",
            "garantia",
            "status",
            "data_entrada",
            "data_saida",
        }
        if ordenar_por not in campos_ordenacao_permitidos:
            raise AtivoErro("Campo de ordenacao invalido.")

        _validar_intervalo_datas(
            filtros.get("data_entrada_inicial"),
            filtros.get("data_entrada_final"),
            campo="data de entrada",
        )
        _validar_intervalo_datas(
            filtros.get("data_saida_inicial"),
            filtros.get("data_saida_final"),
            campo="data de saida",
        )

        tem_garantia = _normalizar_flag_presenca(request.args.get("tem_garantia"))
        tem_nota_fiscal = _normalizar_flag_presenca(request.args.get("tem_nota_fiscal"))
        return filtros, ordenar_por, ordem, tem_garantia, tem_nota_fiscal

    def _filtrar_presenca_documentos(
        ativos: list[Ativo],
        *,
        user_id: int,
        tem_garantia: str | None,
        tem_nota_fiscal: str | None,
    ) -> list[Ativo]:
        """
        Aplica filtros de presença documental usando anexos reais como fonte primária.
        Mantém fallback para campos legados do ativo quando não houver anexos.

        Fluxo:
        1. Se arquivo_service possui método mapear_presenca_documentos (batch),
           usa consulta otimizada que evita N+1 por ativo.
        2. Se arquivo_service NÃO possui esse método (testes antigos, etc),
           cai em fallback que chama listar_arquivos() por ativo — mais lento.
        3. Em ambos casos, pode usar campos legados (nota_fiscal, garantia)
           do ativo como última chance se nenhum anexo real foi encontrado.

        Razão do fallback existir: Compatibilidade com testes antigos que usam
        doubles/fakes sem implementar o método em lote. Em produção, espera-se
        que AtivosArquivoService tenha mapear_presenca_documentos().
        """
        # Evita custo extra de I/O quando o usuário não solicitou filtro documental.
        if tem_garantia is None and tem_nota_fiscal is None:
            return ativos

        # Busca presença documental em lote para evitar N+1 por ativo na listagem.
        # A fonte principal continua sendo anexo real; campos legados entram como fallback.
        ativo_ids = [str(ativo.id_ativo) for ativo in ativos if getattr(ativo, "id_ativo", None)]
        usar_fallback_listagem = False
        if hasattr(arquivo_service, "mapear_presenca_documentos"):
            presenca_por_ativo = arquivo_service.mapear_presenca_documentos(ativo_ids, user_id)
        else:
            # FALLBACK: Compatibilidade para doubles de testes antigos sem o método em lote.
            # Em produção, isso NÃO deve ocorrer. Se ocorrer, é sinal de que o serviço
            # de arquivos está incompleto ou desconfigurável.
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(
                "Filtro documental caindo em fallback legado: arquivo_service não possui "
                "método mapear_presenca_documentos(). Isso pode degradar performance da "
                "listagem de ativos com filtros de presença documental. Verifique se "
                "AtivosArquivoService foi inicializado corretamente."
            )
            presenca_por_ativo = {}
            usar_fallback_listagem = True

        def _tem_documento(ativo: Ativo, tipo_documento: str) -> bool:
            """
            Resolve presença documental priorizando anexos persistidos na tabela ativos_arquivos.
            """
            ativo_id = str(getattr(ativo, "id_ativo", "") or "")
            if usar_fallback_listagem:
                try:
                    anexos = arquivo_service.listar_arquivos(ativo_id, user_id)
                except (AtivoErro, ArquivoNaoEncontrado, ArquivoInvalido, TipoDocumentoInvalido, ValueError, TypeError, KeyError):
                    anexos = []

                possui_anexo_do_tipo = any(
                    (anexo.get("tipo_documento") == tipo_documento)
                    and (anexo.get("nome_original") or "").strip()
                    for anexo in anexos
                )
                if possui_anexo_do_tipo:
                    return True

            presenca_lote = presenca_por_ativo.get(ativo_id, {})
            if bool(presenca_lote.get(tipo_documento)):
                return True

            if tipo_documento == "garantia":
                return bool((getattr(ativo, "garantia", "") or "").strip())

            if tipo_documento == "nota_fiscal":
                return bool((getattr(ativo, "nota_fiscal", "") or "").strip())

            return False

        filtrados = ativos

        if tem_garantia == "sim":
            filtrados = [a for a in filtrados if _tem_documento(a, "garantia")]
        elif tem_garantia == "nao":
            filtrados = [a for a in filtrados if not _tem_documento(a, "garantia")]

        if tem_nota_fiscal == "sim":
            filtrados = [a for a in filtrados if _tem_documento(a, "nota_fiscal")]
        elif tem_nota_fiscal == "nao":
            filtrados = [a for a in filtrados if not _tem_documento(a, "nota_fiscal")]

        return filtrados

    def _buscar_ativos_com_filtros(user_id: int) -> list[Ativo]:
        """
        Resolve listagem com reaproveitamento da API atual e filtros complementares de presença.
        """
        filtros, ordenar_por, ordem, tem_garantia, tem_nota_fiscal = _coletar_filtros_e_ordenacao_da_query()

        if filtros:
            ativos = service.filtrar_ativos(
                user_id=user_id,
                filtros=filtros,
                ordenar_por=ordenar_por,
                ordem=ordem,
            )
        else:
            ativos = service.listar_ativos(user_id)

        return _filtrar_presenca_documentos(
            ativos,
            user_id=user_id,
            tem_garantia=tem_garantia,
            tem_nota_fiscal=tem_nota_fiscal,
        )

    @app.get("/dashboard")
    def dashboard():
        """
        Tela principal do sistema após autenticação.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return redirect(url_for("home"))

        # Mantém o dashboard como visão geral com indicadores leves e atalhos.
        indicadores = {
            "total_ativos": 0,
            "total_em_uso": 0,
            "total_disponivel": 0,
            "total_baixado": 0,
        }
        ativos_preview: list[dict] = []
        try:
            ativos = service.listar_ativos(user_id)
            indicadores["total_ativos"] = len(ativos)
            indicadores["total_em_uso"] = sum(1 for a in ativos if (a.status or "").strip().lower() == "em uso")
            indicadores["total_disponivel"] = sum(1 for a in ativos if (a.status or "").strip().lower() == "disponível")
            indicadores["total_baixado"] = sum(1 for a in ativos if (a.status or "").strip().lower() == "baixado")

            # Exibe apenas uma amostra para manter o dashboard como resumo executivo.
            ativos_ordenados = sorted(
                ativos,
                key=lambda item: ((item.data_entrada or "").strip(), str(item.id_ativo)),
                reverse=True,
            )
            ativos_preview = [_serializar_ativo(ativo) for ativo in ativos_ordenados[:5]]
        except AtivoErro:
            # Em caso de falha de leitura, preserva renderização do dashboard sem interromper UX.
            pass

        return render_template(
            "dashboard.html",
            usuario_email=session.get("user_email"),
            status_validos=STATUS_VALIDOS,
            indicadores=indicadores,
            ativos_preview=ativos_preview,
            show_chrome=True,
        )

    @app.get("/ativos")
    def listar_ativos():
        """
        Lista os ativos do usuário autenticado em formato JSON.
        Aceita parâmetros de filtro opcionais.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error(MSG_SESSAO_EXPIRADA, status=401)

        try:
            ativos = _buscar_ativos_com_filtros(user_id)

            return _json_success(
                "Ativos carregados com sucesso.",
                ativos=[_serializar_ativo(ativo) for ativo in ativos],
            )
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except (ValueError, KeyError, TypeError):
            return _json_error(MSG_ERRO_LISTAR_ATIVOS, status=500)

    @app.post("/ativos")
    # Ordem de segurança: autenticação primeiro (401), CSRF depois (403).
    @require_auth_api()
    @require_csrf()
    def criar_ativo(*, user_id: int):
        """
        Cria um novo ativo usando o contrato mínimo do dashboard.

        Requer usuário autenticado (validado por @require_auth_api()).
        Requer token CSRF válido na requisição (validado pelo decorator @require_csrf()).
        Aceita payloads mínimos sem descricao e categoria — esses campos são
        preenchidos automaticamente pelo backend.

        Implementa proteção contra duplicidade:
        - Detecta requisições duplicadas baseado em hash dos dados principais
        - Retorna ativo já criado se a mesma requisição é recebida novamente em <10s
        - Evita múltiplas criações por cliques repetidos do usuário
        """
        try:
            dados = _request_data()
            ativo = _ativo_do_payload(dados)

            # ========== PROTEÇÃO CONTRA DUPLICIDADE ==========
            # Gera uma chave de deduplicação baseada nos dados principais do ativo.
            # Se a mesma chave foi criada nos últimos 10s, retorna ativo existente.
            chave_dedup = _gerar_chave_dedup(dados, user_id)
            eh_duplicada, id_existente = _verificar_duplicacao(chave_dedup, None)

            if eh_duplicada and id_existente:
                # Duplicação detectada — retorna ativo já criado.
                criado = service.buscar_ativo(id_existente, user_id)
                return _json_success(
                    "Ativo já havia sido cadastrado. Retornando ativo existente.",
                    status=201,
                    ativo=_serializar_ativo(criado),
                )

            # Não é duplicação — procede com criação normal.
            id_gerado = service.criar_ativo(ativo, user_id)

            # Atualiza cache com o ID real gerado.
            _verificar_duplicacao(chave_dedup, id_gerado)

            criado = service.buscar_ativo(id_gerado, user_id)
            return _json_success(
                "Ativo cadastrado com sucesso.",
                status=201,
                ativo=_serializar_ativo(criado),
            )
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except ValueError as erro:
            return _json_error(str(erro), status=400)
        except (KeyError, TypeError):
            return _json_error("Erro inesperado ao cadastrar ativo.", status=500)

    @app.get("/ativos/<id_ativo>")
    def buscar_ativo(id_ativo):
        """
        Retorna um ativo específico em JSON.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error("Sessão expirada. Faça login novamente.", status=401)

        try:
            ativo = service.buscar_ativo(id_ativo, user_id)
            return _json_success("Ativo carregado com sucesso.", ativo=_serializar_ativo(ativo))
        except AtivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except (PermissaoNegada, AtivoErro) as erro:
            return _json_error(str(erro), status=400)
        except (ValueError, KeyError, TypeError):
            return _json_error("Erro inesperado ao consultar ativo.", status=500)

    @app.put("/ativos/<id_ativo>")
    # Ordem de segurança: autenticação primeiro (401), CSRF depois (403).
    @require_auth_api()
    @require_csrf()
    def atualizar_ativo(id_ativo, *, user_id: int):
        """
        Atualiza um ativo específico via fetch.

        Requer usuário autenticado (validado por @require_auth_api()).
        Requer token CSRF válido na requisição (validado pelo decorator @require_csrf()).
        Payload deve conter campos do ativo para atualização.
        """
        dados = _request_data()
        dados_normalizados = _normalizar_payload_atualizacao(dados)

        try:
            ativo = service.atualizar_ativo(id_ativo=id_ativo, dados=dados_normalizados, user_id=user_id)
            return _json_success(
                "Ativo atualizado com sucesso.",
                ativo=_serializar_ativo(ativo),
                resumo_movimentacao=getattr(ativo, "resumo_movimentacao", None),
            )
        except AtivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except (PermissaoNegada, AtivoErro) as erro:
            return _json_error(str(erro), status=400)
        except ValueError as erro:
            return _json_error(str(erro), status=400)
        except (KeyError, TypeError):
            return _json_error("Erro inesperado ao atualizar ativo.", status=500)

    @app.post("/ativos/<id_ativo>/movimentacao/preview")
    # Ordem de segurança: autenticação primeiro (401), CSRF depois (403).
    @require_auth_api()
    @require_csrf()
    def preview_movimentacao_ativo(id_ativo, *, user_id: int):
        """
        Gera prévia estruturada da movimentação sem persistir atualização no banco.

        Requer usuário autenticado (validado por @require_auth_api()).
        Requer token CSRF válido na requisição (validado pelo decorator @require_csrf()).
        Retorna um resumo do que será alterado sem efetuar a mudança.
        """
        dados = _request_data()
        dados_normalizados = _normalizar_payload_atualizacao(dados)

        try:
            preview = service.gerar_preview_atualizacao(
                id_ativo=id_ativo,
                dados=dados_normalizados,
                user_id=user_id,
            )
            return _json_success("Prévia de movimentação gerada com sucesso.", preview=preview)
        except AtivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except (PermissaoNegada, AtivoErro) as erro:
            return _json_error(str(erro), status=400)
        except (ValueError, KeyError, TypeError):
            return _json_error("Erro inesperado ao gerar prévia de movimentação.", status=500)

    @app.post("/ativos/<id_ativo>/movimentacao/confirmar")
    # Ordem de segurança: autenticação primeiro (401), CSRF depois (403).
    @require_auth_api()
    @require_csrf()
    def confirmar_movimentacao_ativo(id_ativo, *, user_id: int):
        """
        Confirma a edição após ajuste operacional no modal e persiste o resultado final.

        Requer usuário autenticado (validado por @require_auth_api()).
        Requer token CSRF válido na requisição (validado pelo decorator @require_csrf()).
        Payload deve conter dados_formulario e ajustes_movimentacao.
        """
        dados = _request_data()
        dados_formulario = dados.get("dados_formulario") or {}
        ajustes_movimentacao = dados.get("ajustes_movimentacao") or {}
        # Na confirmação da movimentação aceitamos payload enxuto para evitar cobrar campos cadastrais.
        dados_normalizados = _normalizar_payload_atualizacao(
            dados_formulario,
            preencher_campos_ausentes=False,
        )

        try:
            dados_finais = service.preparar_dados_confirmacao_movimentacao(
                dados=dados_normalizados,
                ajustes=ajustes_movimentacao,
            )
            ativo = service.atualizar_ativo(id_ativo=id_ativo, dados=dados_finais, user_id=user_id)
            return _json_success(
                "Ativo atualizado com sucesso.",
                ativo=_serializar_ativo(ativo),
                resumo_movimentacao=getattr(ativo, "resumo_movimentacao", None),
            )
        except AtivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except (PermissaoNegada, AtivoErro) as erro:
            return _json_error(str(erro), status=400)
        except (ValueError, KeyError, TypeError):
            return _json_error("Erro inesperado ao confirmar movimentação.", status=500)

    @app.delete("/ativos/<id_ativo>")
    # Ordem de segurança: autenticação primeiro (401), CSRF depois (403).
    @require_auth_api()
    @require_csrf()
    def remover_ativo(id_ativo, *, user_id: int):
        """
        Exclui um ativo específico via fetch.

        Requer usuário autenticado (validado por @require_auth_api()).
        Requer token CSRF válido na requisição (validado pelo decorator @require_csrf()).
        Remove permanentemente o ativo do banco de dados.
        """
        try:
            service.remover_ativo(id_ativo, user_id)
            return _json_success("Ativo removido com sucesso.")
        except AtivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except (PermissaoNegada, AtivoErro) as erro:
            return _json_error(str(erro), status=400)
        except (ValueError, KeyError, TypeError):
            return _json_error("Erro inesperado ao remover ativo.", status=500)

    @app.get("/ativos/lista")
    def listar_ativos_html():
        """
        Renderiza a listagem de ativos com filtros em modal e ações por linha.
        Passa as listas de valores controlados (status, tipos, setores) para renderizar
        filtros como select fields em vez de input text.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return redirect(url_for("home"))

        return render_template(
            "ativos.html",
            usuario_email=session.get("user_email"),
            status_validos=STATUS_VALIDOS,
            tipos_validos=TIPOS_ATIVO_VALIDOS,
            setores_validos=SETORES_VALIDOS,
            show_chrome=True,
        )

    @app.get("/ativos/novo")
    def criar_ativo_html():
        """
        Renderiza tela dedicada de cadastro com seção de documentos vinculados ao ativo.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return redirect(url_for("home"))

        return render_template(
            "novo_ativo.html",
            usuario_email=session.get("user_email"),
            status_validos=STATUS_VALIDOS,
            setores_validos=SETORES_VALIDOS,
            condicoes_validas=CONDICOES_VALIDAS,
            unidades_validas=UNIDADES_VALIDAS,  # Fase 3 Round 2.1
            show_chrome=True,
        )

    @app.get("/ativos/editar/<id_ativo>")
    def editar_ativo_html(id_ativo):
        """
        Renderiza tela dedicada de edição de ativo e gestão de documentos vinculados.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return redirect(url_for("home"))

        return render_template(
            "editar_ativo.html",
            usuario_email=session.get("user_email"),
            status_validos=STATUS_VALIDOS,
            setores_validos=SETORES_VALIDOS,
            condicoes_validas=CONDICOES_VALIDAS,
            unidades_validas=UNIDADES_VALIDAS,  # Fase 3 Round 2.1
            id_ativo=id_ativo,
            read_only=False,
            show_chrome=True,
        )

    @app.get("/ativos/<id_ativo>/editar")
    def editar_ativo_html_alias(id_ativo):
        """
        Alias para manter semantica de rota por recurso sem quebrar URLs legadas.
        """
        return redirect(url_for("editar_ativo_html", id_ativo=id_ativo))

    @app.get("/ativos/visualizar/<id_ativo>")
    def visualizar_ativo_html(id_ativo):
        """
        Mantem compatibilidade do endpoint legado de visualizacao.
        """
        return redirect(url_for("detalhar_ativo_html", id_ativo=id_ativo))

    @app.get("/ativos/detalhes/<id_ativo>")
    def detalhar_ativo_html(id_ativo):
        """
        Renderiza tela de especificacoes completas e documentos do ativo.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return redirect(url_for("home"))

        try:
            ativo = service.buscar_ativo(id_ativo, user_id)
            anexos = arquivo_service.listar_arquivos(id_ativo, user_id)
            ativo_serializado = _serializar_ativo(ativo)

            anexos_nota_fiscal = [a for a in anexos if a.get("tipo_documento") == "nota_fiscal"]
            anexos_garantia = [a for a in anexos if a.get("tipo_documento") == "garantia"]
            anexos_complementares = [a for a in anexos if a.get("tipo_documento") == "outro"]

            # Prioriza documentos realmente vinculados na tabela de anexos.
            # Mantem fallback para os campos legados do ativo, quando existirem.
            ativo_serializado["nota_fiscal"] = (
                anexos_nota_fiscal[0]["nome_original"]
                if anexos_nota_fiscal
                else (ativo_serializado.get("nota_fiscal") or "")
            )
            ativo_serializado["garantia"] = (
                anexos_garantia[0]["nome_original"]
                if anexos_garantia
                else (ativo_serializado.get("garantia") or "")
            )

            return render_template(
                "detalhe_ativo.html",
                usuario_email=session.get("user_email"),
                ativo=ativo_serializado,
                anexos_nota_fiscal=anexos_nota_fiscal,
                anexos_garantia=anexos_garantia,
                anexos_complementares=anexos_complementares,
                show_chrome=True,
            )
        except AtivoNaoEncontrado as erro:
            # Redireciona com flash para evitar JSON bruto em navegação GET.
            flash(str(erro), "danger")
            return redirect(url_for("listar_ativos_html"))
        except (PermissaoNegada, AtivoErro) as erro:
            flash(str(erro), "danger")
            return redirect(url_for("listar_ativos_html"))
        except (OSError, ValueError, TypeError, KeyError):
            flash("Erro ao carregar detalhes do ativo.", "danger")
            return redirect(url_for("listar_ativos_html"))

    @app.get("/ativos/<id_ativo>/detalhes")
    def detalhar_ativo_html_alias(id_ativo):
        """
        Alias para rota de detalhes no padrao por recurso.
        """
        return redirect(url_for("detalhar_ativo_html", id_ativo=id_ativo))

    @app.get("/ativos/<id_ativo>/resumo")
    def obter_resumo_ativo(id_ativo):
        """
        Retorna um resumo composto do ativo em JSON para exibição em modal.
        Inclui informações principais e campos específicos conforme o tipo.

        Controle de visibilidade por perfil:
        - Campos técnicos restritos (AnyDesk, TeamViewer, hostname, serial, código interno)
          são retornados apenas se o usuário for admin.

        Resposta em caso de sucesso:
        {
            "ok": true,
            "resumo": {
                "secao_principal": {...},
                "secao_responsabilidade": {...},
                "secao_ciclo": {...},
                "secao_tecnica": {...},
                "secao_tecnica_restrita": {...}  // Apenas se admin
            }
        }
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error("Sessão expirada.", status=401)

        try:
            ativo = service.buscar_ativo(id_ativo, user_id)
            ativo_serializado = _serializar_ativo(ativo)

            # Obtém perfil do usuário da sessão para controlar visibilidade
            perfil_usuario = session.get("user_perfil")
            eh_admin_user = _eh_admin(perfil_usuario)

            # Compõe o resumo com visibilidade controlada
            resumo = _resumo_ativo_para_modal(ativo_serializado, eh_admin_user)

            return _json_success(
                "Resumo do ativo carregado com sucesso.",
                resumo=resumo,
            )
        except AtivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except (PermissaoNegada, AtivoErro) as erro:
            return _json_error(str(erro), status=400)
        except (OSError, ValueError, TypeError, KeyError) as erro:
            return _json_error("Erro ao carregar resumo do ativo.", status=500)

    @app.post("/ativos/remover/<id_ativo>")
    def remover_ativo_html(id_ativo):
        """
        Mantém compatibilidade com remoções legadas em HTML.
        A interface atual usa DELETE /ativos/<id> via fetch; esta rota é código legado.
        Erros de remoção agora chegam ao usuário via flash para evitar falhas silenciosas.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return redirect(url_for("home"))

        # Compatibiliza rota legada HTML com o mesmo hardening de mutações.
        if not validar_csrf_da_requisicao(request):
            flash("Requisição inválida. Atualize a página e tente novamente.", "danger")
            return redirect(url_for("listar_ativos_html"))

        try:
            service.remover_ativo(id_ativo, user_id)
        except AtivoNaoEncontrado as erro:
            flash(str(erro), "danger")
        except (PermissaoNegada, AtivoErro) as erro:
            flash(str(erro), "danger")

        return redirect(url_for("listar_ativos_html"))

    # ====== ANEXOS (FILES) ======

    @app.post("/ativos/<id_ativo>/anexos")
    # Ordem de segurança: autenticação primeiro (401), CSRF depois (403).
    @require_auth_api()
    @require_csrf()
    def upload_anexo(id_ativo, *, user_id: int):
        """
        Faz upload de um anexo para um ativo.
        Espera: type (nota_fiscal, garantia ou outro), file (arquivo binário)

        Requer usuário autenticado (validado por @require_auth_api()).
        Requer token CSRF válido na requisição (validado pelo decorator @require_csrf()).

        Validações (primeira linha de defesa):
        - Arquivo presente
        - Tipo de documento válido

        Validações no serviço:
        - Tamanho máximo (10 MB)
        - Extensão permitida
        - Mimetype vs extensão (integridade)
        - Arquivo não vazio
        """
        tipo_documento = request.form.get("type", "").strip()
        arquivo = request.files.get("file")

        # Primeira linha de defesa: validar presença de arquivo
        if not arquivo or not arquivo.filename:
            return _json_error("Nenhum arquivo foi enviado.", status=400)

        # Primeira linha de defesa: validar tipo de documento
        tipo_valido = tipo_documento.lower() if tipo_documento else ""
        tipos_permitidos = {"nota_fiscal", "garantia", "outro"}
        if tipo_valido not in tipos_permitidos:
            return _json_error(
                "Tipo de documento inválido. Use: nota_fiscal, garantia ou outro.",
                status=400
            )

        try:
            arquivo_id = arquivo_service.salvar_arquivo(
                ativo_id=id_ativo,
                tipo_documento=tipo_documento,
                arquivo=arquivo,
                user_id=user_id
            )
            return _json_success(
                "Anexo enviado com sucesso.",
                status=201,
                arquivo_id=arquivo_id
            )
        except (TipoDocumentoInvalido, ArquivoInvalido) as erro:
            return _json_error(str(erro), status=400)
        except ArquivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except (StorageBackendError, OSError, ValueError, TypeError) as erro:
            return _json_error(str(erro), status=500)

    @app.get("/ativos/<id_ativo>/anexos")
    def listar_anexos(id_ativo):
        """
        Lista os anexos de um ativo.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error("Sessão expirada. Faça login novamente.", status=401)

        try:
            arquivos = arquivo_service.listar_arquivos(id_ativo, user_id)
            return _json_success(
                "Anexos carregados com sucesso.",
                anexos=[_serializar_arquivo(arquivo) for arquivo in arquivos]
            )
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except (OSError, ValueError, TypeError, KeyError):
            return _json_error("Erro ao listar anexos.", status=500)

    @app.get("/anexos/<int:arquivo_id>/download")
    def download_anexo(arquivo_id):
        """
        Faz download de um anexo específico.

        Comportamento:
        - Se storage local: retorna arquivo via send_file
        - Se storage S3: redireciona para URL assinada de presigned
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error("Sessão expirada. Faça login novamente.", status=401)

        try:
            arquivo = arquivo_service.obter_arquivo(arquivo_id, user_id)
            storage_backend = arquivo_service.storage_backend
            caminho_relativo = arquivo["caminho_arquivo"]

            # Para S3 (presigned URL), redireciona direto para o S3.
            # Para local, carrega arquivo em memória e serve via Flask.
            if isinstance(storage_backend, S3StorageBackend):
                url_assinada = storage_backend.get_download_url(caminho_relativo)
                return redirect(url_assinada)

            # Storage local: carrega arquivo do backend.
            arquivo_bytes = storage_backend.load(caminho_relativo)
            return send_file(
                arquivo_bytes,
                as_attachment=True,
                download_name=arquivo["nome_original"],
                mimetype=arquivo["mime_type"] or "application/octet-stream"
            )

        except ArquivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except (StorageBackendError, OSError, ValueError, TypeError, KeyError) as erro:
            return _json_error(str(erro), status=500)

    @app.delete("/anexos/<int:arquivo_id>")
    # Ordem de segurança: autenticação primeiro (401), CSRF depois (403).
    @require_auth_api()
    @require_csrf()
    def remover_anexo(arquivo_id, *, user_id: int):
        """
        Remove um anexo específico.

        Requer usuário autenticado (validado por @require_auth_api()).
        Requer token CSRF válido na requisição (validado pelo decorator @require_csrf()).
        Remove permanentemente o arquivo do armazenamento e do banco de dados.
        """
        try:
            arquivo_service.remover_arquivo(arquivo_id, user_id)
            return _json_success("Anexo removido com sucesso.")
        except ArquivoNaoEncontrado as erro:
            return _json_error(str(erro), status=404)
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except (OSError, ValueError, TypeError, KeyError):
            return _json_error("Erro ao remover anexo.", status=500)

    # ====== EXPORT / IMPORT ======

    @app.get("/ativos/export/csv")
    def exportar_ativos_csv():
        """
        Exporta a lista de ativos (com filtros opcionais) em formato CSV.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error("Sessão expirada. Faça login novamente.", status=401)

        try:
            ativos = _buscar_ativos_com_filtros(user_id)
            if not ativos:
                return _resposta_sem_registros_exportacao()

            linhas = _linhas_exportacao_enriquecidas(ativos, user_id)

            # Monta CSV em memória
            output = io.StringIO()
            writer = csv.DictWriter(
                output,
                fieldnames=[
                    "id", "tipo", "marca", "modelo", "usuario_responsavel",
                    "departamento", "status", "data_entrada", "data_saida",
                    "nota_fiscal", "garantia"
                ]
            )
            writer.writeheader()
            for linha in linhas:
                writer.writerow(linha)
            
            csv_bytes = io.BytesIO(output.getvalue().encode("utf-8-sig"))
            
            return send_file(
                csv_bytes,
                as_attachment=True,
                download_name=_nome_arquivo_exportacao("csv"),
                mimetype="text/csv"
            )
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except (OSError, ValueError, TypeError, KeyError, csv.Error):
            return _json_error("Erro ao exportar ativos.", status=500)

    @app.get("/ativos/export/xlsx")
    def exportar_ativos_xlsx():
        """
        Exporta a lista de ativos (com filtros opcionais) em formato XLSX real.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error("Sessão expirada. Faça login novamente.", status=401)

        try:
            ativos = _buscar_ativos_com_filtros(user_id)
            if not ativos:
                return _resposta_sem_registros_exportacao()

            xlsx_bytes = _gerar_xlsx_em_memoria(_linhas_exportacao_enriquecidas(ativos, user_id))

            return send_file(
                xlsx_bytes,
                as_attachment=True,
                download_name=_nome_arquivo_exportacao("xlsx"),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except (OSError, ValueError, TypeError, KeyError):
            return _json_error("Erro ao exportar ativos em XLSX.", status=500)

    @app.get("/ativos/export/pdf")
    def exportar_ativos_pdf():
        """
        Exporta a lista de ativos (com filtros opcionais) em formato PDF real.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error("Sessão expirada. Faça login novamente.", status=401)

        try:
            ativos = _buscar_ativos_com_filtros(user_id)
            if not ativos:
                return _resposta_sem_registros_exportacao()

            pdf_bytes = _gerar_pdf_em_memoria(_linhas_exportacao_enriquecidas(ativos, user_id))

            return send_file(
                pdf_bytes,
                as_attachment=True,
                download_name=_nome_arquivo_exportacao("pdf"),
                mimetype="application/pdf",
            )
        except AtivoErro as erro:
            return _json_error(str(erro), status=400)
        except (OSError, ValueError, TypeError, KeyError):
            return _json_error("Erro ao exportar ativos em PDF.", status=500)

    @app.get("/ativos/export/<formato>")
    def exportar_ativos(formato: str):
        """
        Controlador central de exportacao com validacao explicita de formato.
        """
        user_id = _obter_user_id_logado()
        if user_id is None:
            return _json_error("Sessão expirada. Faça login novamente.", status=401)

        formato_normalizado = (formato or "").strip().lower()
        formatos_permitidos = {"csv", "xlsx", "xls", "pdf", "json"}

        if formato_normalizado not in formatos_permitidos:
            return _json_error("Formato de exportação não suportado.", status=400)

        if formato_normalizado in {"xlsx", "xls"}:
            return exportar_ativos_xlsx()

        if formato_normalizado == "pdf":
            return exportar_ativos_pdf()

        if formato_normalizado == "csv":
            return exportar_ativos_csv()

        if formato_normalizado == "json":
            try:
                ativos = _buscar_ativos_com_filtros(user_id)
                linhas = _linhas_exportacao_enriquecidas(ativos, user_id)
                return jsonify(
                    {
                        "ok": True,
                        "formato": "json",
                        "total": len(ativos),
                        "ativos": linhas,
                    }
                )
            except AtivoErro as erro:
                return _json_error(str(erro), status=400)
            except (OSError, ValueError, TypeError, KeyError, csv.Error):
                return _json_error("Erro ao exportar ativos.", status=500)

        return _json_error("Formato de exportação não suportado.", status=400)

    @app.post("/ativos/import/csv")
    # Ordem de segurança: autenticação primeiro (401), CSRF depois (403).
    @require_auth_api()
    @require_csrf()
    def importar_ativos_csv(*, user_id: int):
        """
        Importa ativos a partir de um arquivo CSV.
        Esperado: file (arquivo CSV com headers: id, tipo, marca, modelo, etc)

        Requer usuário autenticado (validado por @require_auth_api()).
        Requer token CSRF válido na requisição (validado pelo decorator @require_csrf()).
        Processa o arquivo CSV linha por linha, criando novos ativos no banco.
        """
        arquivo = request.files.get("file")
        if not arquivo:
            return _json_error("Nenhum arquivo foi enviado.", status=400)

        try:
            # Lê o arquivo CSV
            stream = io.TextIOWrapper(arquivo.stream, encoding="utf-8")
            reader = csv.DictReader(stream)
            
            if not reader.fieldnames:
                return _json_error("Arquivo CSV vazio ou inválido.", status=400)
            
            criados = 0
            erros = []
            
            for idx, row in enumerate(reader, start=2):  # start=2 pois header é linha 1
                try:
                    tipo_csv = row.get("tipo_ativo", row.get("tipo", "")).strip()
                    marca_csv = row.get("marca", "").strip()
                    modelo_csv = row.get("modelo", "").strip()

                    ativo = Ativo(
                        id_ativo=row.get("id", "").strip(),
                        tipo=tipo_csv,
                        tipo_ativo=tipo_csv,
                        marca=marca_csv,
                        modelo=modelo_csv,
                        descricao=row.get("descricao", "").strip() or " ".join(
                            part for part in [tipo_csv, marca_csv, modelo_csv] if part
                        ),
                        categoria=row.get("categoria", "").strip() or tipo_csv,
                        codigo_interno=row.get("codigo_interno", "").strip() or None,
                        serial=row.get("serial", "").strip() or None,
                        condicao=row.get("condicao", "").strip() or None,
                        localizacao=row.get("localizacao", "").strip() or None,
                        setor=row.get("setor", row.get("departamento", "")).strip(),
                        usuario_responsavel=row.get("usuario_responsavel", "").strip() or None,
                        email_responsavel=row.get("email_responsavel", "").strip() or None,
                        departamento=row.get("setor", row.get("departamento", "")).strip(),
                        status=row.get("status", "").strip(),
                        data_entrada=row.get("data_entrada", "").strip(),
                        data_saida=row.get("data_saida", "").strip() or None,
                        data_compra=row.get("data_compra", "").strip() or None,
                        valor=row.get("valor", "").strip() or None,
                        observacoes=row.get("observacoes", "").strip() or None,
                        detalhes_tecnicos=row.get("detalhes_tecnicos", "").strip() or None,
                        processador=row.get("processador", "").strip() or None,
                        ram=row.get("ram", "").strip() or None,
                        armazenamento=row.get("armazenamento", "").strip() or None,
                        sistema_operacional=row.get("sistema_operacional", "").strip() or None,
                        carregador=row.get("carregador", "").strip() or None,
                        teamviewer_id=row.get("teamviewer_id", "").strip() or None,
                        anydesk_id=row.get("anydesk_id", "").strip() or None,
                        nome_equipamento=row.get("nome_equipamento", "").strip() or None,
                        hostname=row.get("hostname", "").strip() or None,
                        imei_1=row.get("imei_1", "").strip() or None,
                        imei_2=row.get("imei_2", "").strip() or None,
                        numero_linha=row.get("numero_linha", "").strip() or None,
                        operadora=row.get("operadora", "").strip() or None,
                        conta_vinculada=row.get("conta_vinculada", "").strip() or None,
                        polegadas=row.get("polegadas", "").strip() or None,
                        resolucao=row.get("resolucao", "").strip() or None,
                        tipo_painel=row.get("tipo_painel", "").strip() or None,
                        entrada_video=row.get("entrada_video", "").strip() or None,
                        fonte_ou_cabo=row.get("fonte_ou_cabo", "").strip() or None,
                        nota_fiscal=row.get("nota_fiscal", "").strip() or None,
                        garantia=row.get("garantia", "").strip() or None,
                    )
                    service.criar_ativo(ativo, user_id)
                    criados += 1
                except AtivoErro as e:
                    erros.append(f"Linha {idx}: {str(e)}")
                except (AttributeError, KeyError, TypeError, ValueError) as e:
                    erros.append(f"Linha {idx}: Erro inesperado - {str(e)}")
            
            msg = f"Importação concluída: {criados} ativo(s) criado(s)."
            if erros:
                msg += f" {len(erros)} erro(s)."
            
            return _json_success(
                msg,
                status=201,
                criados=criados,
                erros=erros if erros else None
            )
        except (OSError, UnicodeError, ValueError, TypeError, KeyError, csv.Error):
            return _json_error("Erro ao processar arquivo CSV.", status=500)
